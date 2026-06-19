import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils  import get_column_letter


_C_HEADER = "0D3347"
_C_ACCENT = "00E5FF"
_C_ALT    = "0A2535"
_C_RED    = "FF3D71"
_C_ORANGE = "FF8C00"
_C_YELLOW = "FFD740"
_C_GREEN  = "A0FF70"
_C_WHITE  = "FFFFFF"

_LEVEL_COLOR = {
    "critical": _C_RED, "high": _C_ORANGE, "medium": _C_YELLOW, "low": _C_GREEN,
}
_STATUS_COLOR = {"failed": _C_RED, "failure": _C_RED, "denied": _C_RED}


def _fill(color: str) -> PatternFill:
    return PatternFill(patternType="solid", fgColor=color)


def _header_font(color: str = _C_ACCENT) -> Font:
    return Font(bold=True, color=color, name="Consolas", size=10)


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)


def _write_header(ws, headers: list, row: int = 1) -> None:
    ws.append(headers)
    for cell in ws[row]:
        cell.font      = _header_font()
        cell.fill      = _fill(_C_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def generate_excel(summary: dict, filename: str) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _sheet_summary(wb, summary, filename)
    _sheet_suspicious(wb, summary)
    _sheet_alerts(wb, summary)
    _sheet_mitre(wb, summary)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _sheet_summary(wb, summary: dict, filename: str) -> None:
    ws = wb.create_sheet("الملخص")
    ws.sheet_view.rightToLeft = True
    _write_header(ws, ["المقياس", "القيمة"])

    rows = [
        ("اسم الملف",              filename),
        ("تاريخ التحليل",          datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("إجمالي السجلات",         summary.get("total_logs", 0)),
        ("الأنشطة المشبوهة",       summary.get("suspicious_events", 0)),
        ("الأنشطة الطبيعية",       summary.get("normal_events", 0)),
        ("مستوى الخطر",            f"{summary.get('risk_score', 0)}%"),
        ("النموذج المستخدم",       summary.get("model_name", "—")),
        ("دقة النموذج (Accuracy)", f"{summary.get('model_accuracy', 0)}%"),
        ("F1 Score",               f"{summary.get('model_f1', 0)}%"),
        ("Precision",              f"{summary.get('model_precision', 0)}%"),
        ("Recall",                 f"{summary.get('model_recall', 0)}%"),
        ("حجم بيانات التدريب",    summary.get("train_size", 0)),
        ("حجم بيانات الاختبار",   summary.get("test_size", 0)),
    ]
    for i, (k, v) in enumerate(rows, 2):
        ws.append([k, str(v)])
        bg = _C_ALT if i % 2 == 0 else _C_HEADER
        for cell in ws[i]:
            cell.fill = _fill(bg)
            cell.font = Font(color=_C_WHITE, name="Consolas", size=10)
    _auto_width(ws)


def _sheet_suspicious(wb, summary: dict) -> None:
    ws = wb.create_sheet("الأنشطة المشبوهة")
    ws.sheet_view.rightToLeft = True

    rows = summary.get("suspicious_table", [])
    if not rows:
        ws.append(["لا توجد أنشطة مشبوهة في هذا التحليل"])
        return

    skip    = {"rule_flag", "ai_flag", "suspicious"}
    headers = [c for c in rows[0].keys() if c not in skip]
    _write_header(ws, headers)

    for i, row in enumerate(rows, 2):
        ws.append([str(row.get(h, "")) for h in headers])
        status = str(row.get("status", "")).lower()
        bg     = _STATUS_COLOR.get(status, _C_ALT if i % 2 == 0 else _C_HEADER)
        for cell in ws[i]:
            cell.fill = _fill(bg)
            cell.font = Font(color=_C_WHITE, name="Consolas", size=9)

    ws.freeze_panes = "A2"
    _auto_width(ws)


def _sheet_alerts(wb, summary: dict) -> None:
    ws = wb.create_sheet("التنبيهات")
    ws.sheet_view.rightToLeft = True

    alerts = summary.get("alerts", [])
    if not alerts:
        ws.append(["لا توجد تنبيهات لهذا التحليل"])
        return

    _write_header(ws, ["المستوى", "العنوان", "الرسالة", "الوقت"])
    for i, a in enumerate(alerts, 2):
        ws.append([a.get("level", ""), a.get("title", ""), a.get("message", ""), a.get("time", "")])
        bg        = _LEVEL_COLOR.get(a.get("level", ""), _C_ALT)
        txt_color = _C_HEADER if bg in (_C_YELLOW, _C_GREEN) else _C_WHITE
        for cell in ws[i]:
            cell.fill = _fill(bg)
            cell.font = Font(bold=True, color=txt_color, name="Consolas", size=9)
    _auto_width(ws)


def _sheet_mitre(wb, summary: dict) -> None:
    ws = wb.create_sheet("MITRE ATT&CK")
    ws.sheet_view.rightToLeft = True

    techniques = summary.get("mitre", {}).get("matched_techniques", [])
    if not techniques:
        ws.append(["لم يتم اكتشاف تقنيات MITRE ATT&CK في هذا التحليل"])
        return

    _write_header(ws, ["المعرّف", "الاسم", "الاسم بالعربي", "التكتيك", "الخطورة", "الوصف"])
    for i, t in enumerate(techniques, 2):
        ws.append([
            t.get("id", ""), t.get("name", ""), t.get("name_ar", ""),
            t.get("tactic_ar", ""), t.get("severity", ""), t.get("description", ""),
        ])
        bg        = _LEVEL_COLOR.get(t.get("severity", ""), _C_ALT)
        txt_color = _C_HEADER if bg in (_C_YELLOW, _C_GREEN) else _C_WHITE
        for cell in ws[i]:
            cell.fill = _fill(bg)
            cell.font = Font(color=txt_color, name="Consolas", size=9)
    _auto_width(ws)
